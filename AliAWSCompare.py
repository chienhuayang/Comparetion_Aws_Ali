from concurrent.futures import ThreadPoolExecutor, as_completed 
import configparser 
import zipfile 
from odps import ODPS 
import pandas as pd 
import boto3 
import time 
import io 
from openpyxl import load_workbook 
from openpyxl.styles import PatternFill, Font 
from openpyxl.utils import get_column_letter 
from openpyxl.utils.dataframe import dataframe_to_rows 
import os 
import pandasql as ps 
import pandas as pd 
import numpy as np 
from datetime import datetime as dt 
from datetime import datetime 
from openpyxl import Workbook 
from decimal import Decimal 
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 
import threading 
 
#Multi-threading版本 
#使用前請先安裝AWS CLI 與 ALI CLI並設置完成 
 
#Global變數  
global file_path ,save_path,record 
file_path = 'Compare_TableList_test.xlsx' #需比對資料表清單的EXCEL檔案路徑 
save_path = 'D:\\Python_Compare\\Result' #比對結果存放路徑 #需事先建立好資料夾 
 
file_lock = threading.Lock() 
 
#AWS query結果會將月份轉為整數，但ALI不會，因此需要將ALI的月份轉為整數，避免不必要的異常 
def remove_leading_zeros(x): 
    if isinstance(x, str): 
        # 檢查字串是否可以轉換為數字 
        try: 
            float(x) 
            # 如果可以轉換為數字，去除開頭的零 
            return str(x).lstrip('0') or '0'  # 如果結果為空字串，返回 '0' 
        except ValueError: 
            # 如果不能轉換為數字，保持原樣 
            return x 
    return x 
 
 
#將ali資料轉換成一般格式(原為decimal) 
def extract_values(dataframe): 
    for col in dataframe.columns:   
        dataframe[col] = dataframe[col].apply( 
            lambda x:  
            float(x) if isinstance(x, Decimal) else   
            (x[1] if isinstance(x, tuple) else None)   
        ) 
        # 開頭為0的資料去0 
        dataframe[col] = dataframe[col].apply(remove_leading_zeros)  
     
    return dataframe 
 
#AWS query結果會將1.0等類似形式的小數轉為整數，此處還原dataframe的數字格式 
def safe_convert(value): 
    try: 
        return float(value) 
    except (ValueError, TypeError): 
        return value 
 
 
#確保 DataFrame 中的數據在進行後續處理時具有一致的數據類型 
def convert_values(dataframe): 
    for col in dataframe.columns: 
        dataframe[col] = dataframe[col].apply( 
            lambda x:  
            float(x) if isinstance(x, Decimal)  
            else (str(x) if isinstance(x, tuple) else x) 
        ) 
    return dataframe 
 
 
def aws_sql_query(table, query): 
    # 判斷 table 最後更新時間 
    df = pd.DataFrame() # 初始化 DataFrame 
 
    if query == 1: 
        glue_client = boto3.client('glue') 
        response = glue_client.get_table(DatabaseName=('ads' if table[0:3].lower()=='ads' else 'cdm'), Name=table) # 判斷資料表位置 
        last_updated_time = response['Table']['UpdateTime'] 
        last_updated_time = last_updated_time.strftime('%Y-%m-%d %H:%M:%S') 
        df = pd.DataFrame({'table_name': table, 'last_update_time': last_updated_time}, index=[0]) 
        return df 
 
    else: 
        athena_client = boto3.client('athena') 
        s3_client = boto3.client('s3') 
        output_location = "s3://ap-sg-bk/dataplatform/query_rsult/" 
 
        # 執行查詢 
        response = athena_client.start_query_execution( 
            QueryString=query, 
            QueryExecutionContext={'Database': ('ads' if table[0:3]=='ads' else 'cdm')}, 
            ResultConfiguration={'OutputLocation': output_location} 
        ) 
 
        # 獲取查詢執行 ID 
        query_execution_id = response['QueryExecutionId'] 
 
        # 等待查詢完成 
        while True: 
            query_status = athena_client.get_query_execution(QueryExecutionId=query_execution_id) 
            status = query_status['QueryExecution']['Status']['State'] 
             
            if status == 'SUCCEEDED': 
                print('aws query done') 
                print() 
                break 
            elif status in ['FAILED', 'CANCELLED']: 
                print(f'Query failed or was cancelled: {status}') 
                return pd.DataFrame({'Query failed or was cancelled': [111]}) 
            else: 
                print('Query is still running...') 
                time.sleep(2) 
        # 獲取 S3 結果位置
        s3_output = query_status['QueryExecution']['ResultConfiguration']['OutputLocation'] 
        bucket_name = s3_output.split('/')[2] 
        file_key = '/'.join(s3_output.split('/')[3:]) 
        # 從 S3 下載結果並轉為 DataFrame 
        response = s3_client.get_object(Bucket = bucket_name, Key=file_key) 
        file_content = response['Body'].read().decode('utf-8') 
        df = pd.read_csv(io.StringIO(file_content),low_memory=False,na_values=[],keep_default_na=False) 
        for col in df.columns: 
            # 檢查列是否只包含數字（包括小數點和負號） 
            if df[col].astype(str).str.match(r'^-?\d*\.?\d+$').all(): 
                df[col] = pd.to_numeric(df[col], errors='coerce') 
            else: 
                # 如果列包含非數字字符，保持原樣 
                pass 
        # print(df.head(10)) #test 
        df = df.map(safe_convert) 
        print('Processing Done...') 
        # print(df.head(10)) #test 
        # 輸出 DataFrame 
 
        del file_content, response 
 
        return df 
 
#讀取aliyun的config檔案 
def load_aliyun_config(profile='default'): 
    c_file_path = os.path.expanduser('~/.aliyun/config') 
    if not os.path.exists(c_file_path): 
        raise FileNotFoundError(f"Config file not found: {c_file_path}") 
 
    config = configparser.ConfigParser() 
    config.read(c_file_path) 
 
    print(f"Sections found in config file: {config.sections()}")  # 調試信息 
 
    if profile not in config.sections(): 
        raise ValueError(f"Profile {profile} not found in config file") 
 
    a_i= config[profile]['a_i'] 
    a_s = config[profile]['a_s'] 
    region_id = config[profile].get('region_id', 'cn-hangzhou')  # 默認區域 
    return a_i, a_s, region_id 
 
#Aliyun query 
def ali_sql_query(table_name, query, key_column=None): 
    a_i, a_s, region_id = load_aliyun_config() #使用設置好的config檔案登入
    project = 'dp_ads' if table_name[0:3].lower() == 'ads' else 'dp_cdm' 
    endpoint = 'http://service.ap-southeast-1.maxcompute.aliyun.com/api'  # 新加坡 Endpoint 
     
    odps = ODPS(a_i, a_s, project, endpoint) 
    print('Ali connect successfully') 
 
    if query == 1: 
        print('Ali querying row count...') 
        table = odps.get_table(table_name) 
        last_modified_time = table.last_data_modified_time 
        formatted_time = datetime.strftime(last_modified_time, '%Y-%m-%d %H:%M:%S') 
        df = pd.DataFrame({'table_name': table_name, 'last_update_time': formatted_time}, index=[0]) 
        return df 
    else: 
        try: 
            # 確認 query 是否正確 
            if pd.isna(query) or query.strip() == '': 
                raise ValueError('Query is empty or NaN') 
            print('Querying Ali the whole table...') 
            with odps.execute_sql(query).open_reader() as reader: 
                rawdata = pd.DataFrame(reader, columns=[col.name for col in reader.schema.columns]) 
                print('Ali query done') 
                print() 
                df = extract_values(rawdata) 
                if key_column: 
                    df = df.sort_values(by=key_column) 
                df = df.map(safe_convert) 
            return df 
        except Exception as e: 
            print(f'Error executing query: {e}') 
            return pd.DataFrame() 
         
#提供快速存檔，較to_excel快速；並且可以設定存檔最大檔案大小 
def fast_excel_write(data, filename, sheet_name, max_size_mb=5): 
    max_size = max_size_mb * 1024 * 1024  # 轉換為字節 
     
    # 確保目錄存在 
    os.makedirs(os.path.dirname(filename), exist_ok=True) 
     
    if os.path.exists(filename): 
        try: 
            # 如果文件已经存在，則打開他並添加新工作表 
            wb = load_workbook(filename) 
            # 如果同名工作表存在，删除它 
            if sheet_name in wb.sheetnames: 
                del wb[sheet_name] 
                print(f'Removing existing sheet "{sheet_name}".') 
            ws = wb.create_sheet(sheet_name) 
            print(f'Adding new sheet "{sheet_name}" to existing file.') 
        except Exception as e: 
            print(f'Error loading workbook: {e}') 
            # 如果文件無法打開，建一个新的工作簿 
            wb = Workbook() 
            ws = wb.active 
            ws.title = sheet_name
print(f'Creating new file with sheet "{sheet_name}".') 
    else: 
        # 如果文件不存在，則創建一个新的工作簿 
        wb = Workbook() 
        ws = wb.active 
        ws.title = sheet_name 
        print(f'Creating new file with sheet "{sheet_name}".') 
     
    # Write the header 
    ws.append(data.columns.tolist()) 
     
    # Write the data in chunks to avoid memory issues #分批寫入 
    chunk_size = 1000 
    rows_written = 0 
    for i in range(0, len(data), chunk_size): 
        chunk = data.iloc[i:i + chunk_size] 
        for row in chunk.itertuples(index=False, name=None): 
            ws.append(row) 
            rows_written += 1 
         
        # 每次寫入後檢查文件大小 
        with file_lock: 
            wb.save(filename) 
        current_size = os.path.getsize(filename) 
         
        if current_size > max_size: 
            print(f"File size limit reached. Wrote {rows_written} rows.") 
            return rows_written 
     
    try: 
        # 對工作簿進行操作 
        with file_lock: 
            wb.save(filename) 
    except Exception as e: 
        print(f"保存檔案時發生錯誤: {str(e)}") 
    finally: 
        # 確保資源被釋放 
        wb.close() 
    print(f"Finished writing {rows_written} rows.") 
 
#比較兩個雲端資料表的資料筆數和更新時間，確保它們的一致性。 
def data_quality_check(table_name, df_ali, df_aws,query_aws,query_ali,sort_query_ali,record,result_path,owner): 
 
    #可優化:將此流程合併至第一次query中，減少Query次數  #已完成  
    isequal = True 
    print('Now in data_quality_check...') 
 
    aws_df = aws_sql_query(table_name,1) 
    ali_df = ali_sql_query(table_name,1) 
 
    try: 
        aws_count = df_aws.shape[0] if not df_aws.empty else '0' #計算aws資料筆數 
     
    except Exception as e: 
        print(f'Error fetching AWS count: {e}') 
        aws_count = '0' 
 
    try: 
        ali_count = df_ali.shape[0] if not df_ali.empty else '0' #計算ali資料筆數 
    except Exception as e: 
        print(f'Error fetching Ali count: {e}') 
        ali_count = '0' 
 
    # 檢查資料筆數是否相等，不相等或都是0則標記為異常且isequal紀錄false(不進入compare) 
    if ali_count != aws_count : 
        print(f'{table_name} Data row count NOT equal.') 
        new_record = pd.DataFrame({'table_name': [table_name], 'status': ['資料筆數異常'],'owner':[owner]}) 
        record = pd.concat([record, new_record], ignore_index=True) 
        isequal = False 
    elif ali_count == '0' and aws_count == '0': 
        new_record = pd.DataFrame({'table_name': [table_name], 'status': ['資料筆數為0'],'owner':[owner]}) 
        record = pd.concat([record, new_record], ignore_index=True) 
        # 如果兩個表中都沒有數據，則不進行比較且標示異常回報false 
        print(f'{table_name} No data in both tables.') 
        isequal = False 
    else: 
        print('Data row count equal.') 
 
     
    sql = fr''' 
    select  
    a.table_name, 
    '{aws_count}' as aws_count, 
 
    '{ali_count}' as ali_count, 
    a.last_update_time as aws_update_time, 
    b.last_update_time as ali_update_time 
    from aws_df a 
    left join ali_df b 
    on a.table_name = b.table_name; 
    ''' 
 
    df = ps.sqldf(sql) 
 
    #將資料品質檢查結果寫入excel 
    dfresult = pd.DataFrame({'table_name': [],'aws_count':[],'ali_count':[],'aws_update_time':[],'ali_update_time':[]}) 
    dfresult = pd.concat([dfresult,df]) 
 
    #20250117 - 將ali和aws的query語法寫入excel，確保資料品質檢查的可追蹤性 
    aws_ds = query_aws.split('ds = ')[1].split('and')[0].strip().replace("'",'') if 'ds = ' in query_aws else '' 
    ali_ds = query_ali.split('ds = ')[1].split('and')[0].strip().replace("'",'') if 'ds = ' in query_ali else '' 
    query_aws = query_aws.split('and', 1)[1].strip() if 'and' in query_aws.lower() else '' 
 
    dfresult['aws_ds'] = aws_ds 
    dfresult['ali_ds'] = ali_ds 
    dfresult['比對條件'] = query_aws 
    dfresult['額外阿里條件'] = sort_query_ali 
 
    fast_excel_write(dfresult , result_path, '資料品質檢查') 
 
    print('Data quality check done.') 
    print() 
 
    return isequal,record 
 
#比較兩個雲端資料表的資料 
def compare(dfali,dfaws,ignore_columns,ignore_asx_count,des,distinct,record,result_path,owner): 
 
    print("Now in compare session...") 
    columnsali = (dfali.columns).to_numpy() #將ali的欄位擷取出並轉為numpy
    columnsaws = (dfaws.columns).to_numpy() #將aws的欄位擷取出並轉為numpy 
 
    #忽略不需要的欄位 
    for i in ignore_columns: 
        columnsali = np.delete(columnsali,np.where(columnsali == i))  
        columnsaws = np.delete(columnsaws,np.where(columnsaws == i)) #刪除忽略欄位 
 
    columnsali = np.sort(columnsali) 
    columnsaws = np.sort(columnsaws) 
 
    sql_query = fr'''select {distinct}''' 
    sql_queryaws = fr'''select {distinct}'''    #query語法 
    oncolumns = '' 
    oncolumnsaws = '' 
 
    index = 0 
    while index < np.count_nonzero(columnsali): #確保欄位數量一致 
 
        try: 
            sql_query = sql_query + ' ' +fr'''  
            trim(case when substr(cast("{columnsali[index]}" as varchar), -2, 2) = '.0' then replace(cast("{columnsali[index]}" as varchar),'.0','') 
            else cast("{columnsali[index]}" as varchar) end) "{columnsali[index]}",''' 
            #aws 
            sql_queryaws = sql_queryaws + ' ' +fr'''  
            trim(case when substr(cast("{columnsaws[index]}" as varchar), -2, 2) = '.0' then replace(cast("{columnsaws[index]}" as varchar),'.0','') 
            else cast("{columnsaws[index]}" as varchar) end) "{columnsaws[index]}",'''   
     
            index += 1 
 
        except Exception as e: 
            print(f'Error executing query: {e}') 
            print("Count of columns in Ali and AWS are not equal.") 
            return pd.DataFrame() 
 
    index = 0 
 
    try: 
        ignore_asx_count = int(ignore_asx_count) 
    except: 
        pass 
 
    #判斷變數是否為數字，不適的話則按照提供的key欄位排序 
    if isinstance(ignore_asx_count, (int, float)): 
 
        while index < np.count_nonzero(columnsali)-ignore_asx_count: 
 
 
            #ali排序 
            oncolumns = oncolumns + fr'''case when substr( 
            lower(trim(COALESCE(case when replace(cast("{columnsali[index]}" as varchar),'\\N','') = '' then '' 
            else replace(cast("{columnsali[index]}" as varchar),'\\N','') end,''))),1,2) = '0.' then  
            cast(round(cast( 
            lower(trim(COALESCE(case when replace(cast("{columnsali[index]}" as varchar),'\\N','') = '' then '' 
            else replace(cast("{columnsali[index]}" as varchar),'\\N','') end,''))) as real),2) as varchar) 
            else  
            lower(trim(COALESCE(case when replace(cast("{columnsali[index]}" as varchar),'\\N','') = '' then '' 
            else replace(cast("{columnsali[index]}" as varchar),'\\N','') end,''))) end || ''' 
 
 
            #aws排序 
            oncolumnsaws = oncolumnsaws + fr'''case when substr( 
            lower(trim(COALESCE(cast("{columnsaws[index]}" as varchar),''))),1,2) = '0.' then  
            cast(round(cast( 
            lower(trim(COALESCE(cast("{columnsaws[index]}" as varchar),''))) as real),2) as varchar) 
            else  
            lower(trim(COALESCE(cast("{columnsaws[index]}" as varchar),''))) end || ''' 
 
            index += 1 
    else: 
        oncolumns = fr''' trim(cast({ignore_asx_count} as varchar)) || ''' 
        oncolumnsaws = fr''' trim(cast({ignore_asx_count} as varchar)) || ''' 
 
    # 在插入資料之前，先進行轉換避免decimal格式的數據 
    dfali = convert_values(dfali) 
    dfaws = convert_values(dfaws) 
 
    #ali設定統整 
    sql_queryali = fr''' 
                        with alldata as ( 
                        {sql_query} {oncolumns[:-3]} as key, 
                        'ali' as datatype 
                        from dfali) 
                        select * from alldata 
                        ;  ''' 
 
    #aws設定統整 
    sql_queryaws = fr'''  
    with alldata as ( 
                        {sql_queryaws}  {oncolumnsaws[:-3]} as key, 
                        'aws' as datatype 
                        from dfaws) 
                        select * from alldata 
                        ;''' 
 
    alitotaldata = ps.sqldf(sql_queryali) #select統整好的資料 
    alitotaldata = alitotaldata.sort_values(by='key',ascending=des) #以key降冪排序 
    alitotaldata = alitotaldata.reset_index(drop=True) #重設index 
    alitotaldata = alitotaldata.reset_index() 
 
    awstotaldata = ps.sqldf(sql_queryaws) 
    awstotaldata = awstotaldata.
    sort_values(by='key',ascending=des) 
    awstotaldata = awstotaldata.reset_index(drop=True) 
    awstotaldata = awstotaldata.reset_index() 
 
    #合併ali和aws的資料 
    all_merge_sql = '''select * from alitotaldata union all select * from awstotaldata;'''  
    all_merge = ps.sqldf(all_merge_sql) 
    all_merge = all_merge.sort_values(by=['index','datatype'],ascending=True) #以index和datatype排序 
    all_merge['index'] = all_merge['index'].astype(str) #index轉為字串 
 
    if all_merge.shape[0] >= 200000 and len(all_merge) > 80 : 
        all_merge = all_merge.head(40000) 
    elif all_merge.shape[0] >= 200000 and len(all_merge) > 50 and len(all_merge) < 80: 
        all_merge = all_merge.head(80000) 
 
    # 新增：比對相同 index 的數據並只保留有差異的行 
    all_merge = compare_rows(all_merge) 
 
    # 檢查 all_merge 是否為空 
    if all_merge.empty: 
        print("No differences found. Saving empty dataframe.") 
        all_merge = pd.DataFrame(columns=['無異常資料']) 
 
    #合併結果存入excel 
    fast_excel_write(all_merge, result_path, '差異交叉顯示') 
 
    print("Comparison done.") 
    print() 
 
    return all_merge.head(10),record #返回前10筆資料(能確認有資料即可) 
 
def preprocess_data(df): 
    # 將 None 和空白字符串統一轉換為 np.nan 
    df = df.map(lambda x: np.nan if pd.isna(x) or (isinstance(x, str) and x.strip() == '') else x) 
     
    # 將所有字符串類型的列去除首尾空白 
    for col in df.select_dtypes(include=['object']): 
        df[col] = df[col].str.strip() 
     
    return df 
 
def compare_rows(df): 
    # 預處理數據 
    df = preprocess_data(df) 
 
    # 檢查 'index' 列是否存在 
    if 'index' not in df.columns: 
        print("Warning: 'index' column not found. Using DataFrame index instead.") 
        df['index'] = df.index.astype(str) 
 
    # 將 DataFrame 按 index 分組 
    grouped = df.groupby('index') 
     
    # 用於存儲有差異的行 
    diff_rows = [] 
     
    columns_to_compare = [col for col in df.columns if col not in ['datatype', 'index','key']] 
 
    for name, group in grouped: 
        if len(group) == 2:  # 確保每組只有兩行數據 
            row1 = group.iloc[0] 
            row2 = group.iloc[1] 
             
            # 檢查是否有任何差異 
            is_different = False 
            for col in columns_to_compare: 
                # 使用 pd.isna 來比較，這樣可以正確處理 np.nan 
                if not pd.isna(row1[col]) or not pd.isna(row2[col]): 
                    if row1[col] != row2[col]: 
                        is_different = True 
                        break 
             
            if is_different: 
                diff_rows.append(row1) 
                diff_rows.append(row2) 
        elif len(group) > 2: 
            print(f"Warning: More than 2 rows found for index {name}. Skipping this group.") 
 
    # 創建新的 DataFrame，只包含有差異的行 
    if diff_rows: 
        result_df = pd.DataFrame(diff_rows) 
         
        # 檢查 'index' 和 'datatype' 列是否存在 
        sort_columns = [] 
        if 'index' in result_df.columns: 
            sort_columns.append('index') 
        if 'datatype' in result_df.columns: 
            sort_columns.append('datatype') 
         
        # 按原始順序排序 
        if sort_columns: 
            result_df = result_df.sort_values(by=sort_columns) 
         
        print(f"Found {len(result_df)} rows with differences.") 
        return result_df 
    else: 
        print("No differences found between rows.") 
        return pd.DataFrame()  # 返回空的 DataFrame 
 
#不一致資料標記 
#taiga highlight python script 
def apply_highlight(_ws, _highlight_df): 
    # 定義黑底紅黃字 
    fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid') 
    a_font, b_font = Font(color='FF0000'), Font(color='FFFF00') 
 
    for row in _highlight_df.itertuples(): 
        idx = row.Index + 2  # +2 是因為openpyxl的行索引從1開始，且有標題行 
        _font = a_font if _ws[f'A{idx}'].value == 'ALI' else b_font 
        for col_idx, cell_value in enumerate(row[1:], 1):  # 從第1列開始 
            if cell_value:  # 如果需要高亮 
                col_letter = get_column_letter(col_idx) 
                _ws[f'{col_letter}{idx}'].fill = fill 
                _ws[f'{col_letter}{idx}'].font = _font 
 

#比較結果標記
def compare_highlight(result_path): 
    target = result_path 
    output = result_path 
    excel_file = pd.ExcelFile(target) 
    sheets = [name for name in excel_file.sheet_names] 
 
    wb = load_workbook(target) 
    for index, sheet in enumerate(sheets, start=1): 
        _log_prefix = f'[{index:02d}/{len(sheets):02d}]' 
        print(f'{_log_prefix} Processing {sheet}...') 
        df = pd.read_excel(target, engine='openpyxl', dtype=str, keep_default_na=False, sheet_name=sheet).dropna(how='all') 
         
        if df.empty or len(df.columns) < 2: 
            print(f'{_log_prefix} Skipping empty or invalid sheet: {sheet}') 
            continue 
         
        df_sorted = df.sort_values(by=df.columns[1], ascending=True) 
        other_columns = [col for col in df.columns if col != '資料位置'] 
 
        highlight_df = pd.DataFrame(False, index=df_sorted.index, columns=df_sorted.columns) 
         
        try: 
            for name, group in df_sorted.groupby('index'): 
                if len(group) > 1: 
                    # 如果同一個資料位置有多筆資料，則判斷每個欄位的值 
                    for col in other_columns: 
                        # 只要有值不一樣，標記需要 Highlight 的 cells 
                        if not group[col].eq(group[col].iloc[0]).all(): 
                            highlight_df.loc[group.index, col] = True 
 
            ws = wb[sheet] 
            print(f'{_log_prefix} Applying highlight...') 
            apply_highlight(ws, highlight_df) 
        except Exception as e: 
            print(f'Error processing sheet {sheet}: {e}') 
            pass 
 
    print('Saving...') 
    try: 
        wb.save(output) 
        print('Done.') 
    except PermissionError: 
        print('請確認檔案未開啟，並且有寫入權限') 
 
 
#寄送結果 
def send_email_with_attachments(recipient, attachment_dir): 
     
    # Create the email 
    to_list = [f'{recipient}@xxxx.com'] 
    msg = MIMEMultipart() 
    msg['From'] = 'BIS@xxxx.com' 
    msg['To'] = ','.join(to_list) 
    msg['Subject'] = 'Compare Result' 
    cc_list = ['Bob@xxxx.com'] 
    msg['CC'] = ','.join(cc_list)   # Alternative way to add CC recipients 
    body = 'Sent from Python script.' 
 
    # Attach the body with the msg instance 
    msg.attach(MIMEText(body, 'plain')) 
    # Create a zip file of the specified directory 
    zip_filename = f'{recipient}_result.zip' 
    with zipfile.ZipFile(zip_filename, 'w') as zipf: 
        for root, dirs, files in os.walk(attachment_dir): 
            for file in files: 
                zipf.write(os.path.join(root, file), os.path.relpath(os.path.join(root, file), attachment_dir)) 
 
    # Attach the zip file 
    with open(zip_filename, "rb") as attachment: 
        part = MIMEBase('application', 'octet-stream') 
        part.set_payload(attachment.read()) 
        encoders.encode_base64(part) 
        part.add_header('Content-Disposition', f"attachment; filename= {zip_filename}") 
        msg.attach(part) 
         
    try: 
        smtp = smtplib.SMTP('xxxx.com.tw', 25) 
        smtp.ehlo() 
        smtp.starttls() 
        recipients = [msg['To']] + cc_list 
        smtp.sendmail(msg['From'], recipients, msg.as_string()) 
        smtp.quit() 
        print(f"{recipient} 郵件傳送成功!") 
    except Exception as e: 
        print(f"{recipient} 郵件傳送失敗: {e}") 
        msg['Subject'] = 'Compare Result' 
        body = f'{recipient} 郵件傳送失敗: {e} . Sent from Python script.' 
        try: 
            smtp = smtplib.SMTP('xxxx.com.tw', 25) 
            smtp.ehlo() 
            smtp.starttls() 
            smtp.sendmail(msg['From'], recipients, msg.as_string()) 
            smtp.quit() 
        except Exception as e: 
            print(f"{recipient} 郵件傳送失敗: {e}") 
         
 
    # Delete the zip file after sending the email 
    if os.path.exists(zip_filename): 
        os.remove(zip_filename) 
     
 
# 編寫query指令 : 檢查 table 是否為 view且有無額外阿里條件 
def query_maker(table_name,sort_query_ali,sort_query_aws): 
 
    # Query文字，含判斷table所在資料庫位置及where(不含條件)
    query_ali = fr'''select * from {('db_a' if table_name[0:3]=='abc' else 'db_c')}.{table_name} where ''' 
    query_aws = fr'''select * from {('a' if table_name[0:3]=='abc' else 'c')}.{table_name} where '''    

    return query_aws,query_ali 
 
# 執行 
def execute(ds,table_name,sort_query_ali,sort_query_aws,ignore_column,record,result_path,owner): 
 
    # 忽略欄位轉為list 
    if pd.isna(ignore_column) or not isinstance(ignore_column, str): # 確保 ignore_column 是字符串並且不為 None 
        ignore_columns = [] 
    else: 
        ignore_columns = ignore_column.split(',') # 以逗號分隔 ignore_column 字符串 
     
    # 編寫query指令 : 檢查 table 是否為 view且有無額外阿里條件 
    query_aws,query_ali = query_maker(table_name,sort_query_ali,sort_query_aws)  
     
    #取得Dataframe 
    aws_df = aws_sql_query(table_name,query_aws) 
    ali_df = ali_sql_query(table_name,query_ali) 
 
    #資料品質檢查 
    print('Ready for data_quality_check...') 
    isequal,record = data_quality_check(table_name,ali_df,aws_df,query_aws,query_ali,sort_query_ali,record,result_path,owner) 
 
    if isequal == False: 
        return pd.DataFrame({'資料數量不對稱或為0無法比對': [111]}),isequal,record 
    else: 
        result,new_record = compare(ali_df,aws_df,ignore_columns,0,True,'',record,result_path,owner) 
 
    if result.empty: 
        print(table_name, '無異常') 
        new_record = pd.DataFrame({'table_name': [table_name], 'status': ['無異常'],'owner':[owner]}) 
        record = pd.concat([record, new_record], ignore_index=True) 
        print(f'{table_name} 無異') 
        print('-----------------------------------') 
        new_result_path = save_path + fr'\{owner}\比對無異常_Compare_Result_{table_name}_{(dt.now().strftime("%Y%m%d"))}.xlsx' 
        if os.path.exists(new_result_path): 
            os.remove(new_result_path) 
        os.rename(result_path, new_result_path) 
    else : 
        compare_highlight(result_path) 
        print(table_name, '有異常') 
        new_record = pd.DataFrame({'table_name': [table_name], 'status': ['資料比對異常'],'owner':[owner]}) 
        record = pd.concat([record, new_record], ignore_index=True) 
        print(f'{table_name} 有異') 
        print('-----------------------------------') 
         
    return result,isequal,record 
 
# 封裝成可以並行處理的函數 
def parallel_execute(args):
    ds, table_name, sort_query_ali, sort_query_aws, ignore_column, record, result_path,owner = args 
    return execute(ds, table_name, sort_query_ali, sort_query_aws, ignore_column, record,result_path,owner) 
 
if name == '__main__':     
    ds = (dt.now() - pd.Timedelta(days=1)).strftime('%Y%m%d') #ds自動抓取前1天日期 
 
    # 讀取 Excel 文件 
    try: 
        df = pd.read_excel(file_path)   #讀取比對清單EXCEL檔案  
    except: 
        print('請確認檔案路徑是否正確') 
        exit(1) 
 
    #創建指定位置第一層資料夾 
    if not os.path.isdir(save_path): 
        os.mkdir(save_path) 
     
    #創建指定資料夾中子資料夾分類不同查詢區間資料 
    if time.time() < time.mktime(time.strptime(f'{dt.now().strftime("%Y-%m-%d")} 15:59:59', '%Y-%m-%d %H:%M:%S')): 
        save_path = save_path + fr'\ds區間-{ds}-上午' 
    else: 
        save_path = save_path + fr'\ds區間-{ds}-下午'   
 
    #確認存取位置存在 
    if not os.path.isdir(save_path): 
        os.mkdir(save_path) 
 
    # 初始化變數 
    table_name = '' 
    sort_query_ali = '' 
    sort_query_aws = '' 
    ignore_column = '' 
    owner = '' 
    default_ignore = 'stat_date,dw_ins_time,ds' # 預設忽略欄位 
    record = pd.DataFrame({'table_name': [],'status':[],'owner':[]}) 
    new_record = pd.DataFrame({'table_name': [],'status':[],'owner':[]}) 
    owner_list = [] 
    tasks = [] 
 
    # 讀取 Excel 中的欄位並將資料回傳至變數 
    for index, row in df.iterrows(): 
        previous_owner = owner 
        table_name = row['table name'] 
        sort_query_aws = row['共用驗證範圍的 where 條件'] 
        sort_query_ali = row['額外阿里的 where 條件'] 
        ignore_column = row['自訂排除欄位'] 
        owner = row['Owner'] 
 
        # 確保 ignore_column 是字符串並且不為 None 
        if pd.isna(ignore_column) or not isinstance(ignore_column, str): 
            ignore_column = default_ignore 
        else: 
            ignore_column = ignore_column + ',' + default_ignore 
 
        # 打印變數值 
        print(f'Table Name: {table_name}') 
        print(f'Sort Query Ali: {sort_query_ali}') 
        print(f'Sort Query AWS: {sort_query_aws}') 
        print(f'Ignore Column: {ignore_column}') 
        print(f'Owner: {owner}') 
        print() 
 
        # 判斷 table 是否比對完成，若空值則完成則退出 
        if table_name == '' or pd.isna(table_name): 
            print('All tables have been insert in tasks') 
        else:             
            #創建結果路徑 
            result_path = save_path + fr'\{owner}\Compare_Result_{table_name}_{(dt.now().strftime("%Y%m%d"))}.xlsx' 
            print('task added')   
            tasks.append((ds, table_name, sort_query_ali, sort_query_aws, ignore_column, record, result_path,owner)) 
            owner_list.append(owner) 
 
    # 使用 ThreadPoolExecutor 來並行運行tasks，最多 3 個執行緒 
    print('Start comparing-----------------------------------') 
    with ThreadPoolExecutor(max_workers=3) as executor: 
        futures = [executor.submit(parallel_execute, task) for task in tasks] 
        for future in as_completed(futures): 
            result,isequal,new_record = future.result() 
            record = pd.concat([new_record, record], ignore_index=True) 
    print('End comparing-----------------------------------') 
 
     
 
    onwer_list = set(owner_list) 
    for owner in onwer_list: 
        print(f'Sending email to {owner}...') 
        fast_excel_write(record, save_path+ fr'\{owner}\紀錄總結.xlsx', '檢查紀錄') 
        #send_email_with_attachments(owner, save_path+ fr'\{owner}') 
        print('Email sent.') 
        print('-----------------------------------') 
         
    print('All tables have been compared')
